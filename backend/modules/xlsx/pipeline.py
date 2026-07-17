"""编排 XLSX 提取、批量翻译、安全回退、结构校验与原子落盘。"""

import os
import re
import tempfile
from collections import Counter
from collections.abc import Sequence
from dataclasses import dataclass, field
from html import unescape
from pathlib import Path

from openpyxl import load_workbook

from modules.translation.batch_fanout import (
    BatchFanoutOutcome,
    run_batch_fanout,
    translator_per_job_concurrency,
)
from modules.translation.contracts import (
    BatchTranslator,
    TranslationBatchResult,
    TranslationProgress,
    TranslationProgressReporter,
    TranslationRequest,
    TranslationResult,
)
from modules.translation.token_groups import group_by_token_limit

from .entities import XlsxCellSegment
from .extractor import XlsxExtractor
from .formatter import XlsxFormatter

_SAFE_LANGUAGE_RE = re.compile(r"[^A-Za-z0-9_-]+")
_SPAN_RE = re.compile(r"^<span>(.*?)</span>$", re.DOTALL)


@dataclass(slots=True)
class _TranslationStats:
    """累计 XLSX cell 的处理、成功、回退与 warning 数。"""

    processed_segments: int = 0
    translated_segments: int = 0
    fallback_segments: int = 0
    warning_codes: set[str] = field(default_factory=set)


@dataclass(frozen=True, slots=True)
class _GroupTranslation:
    """保存一个 cell group 已经安全收敛的结果。"""

    items: tuple[tuple[tuple[int, str], str], ...]
    translated_segments: int
    fallback_segments: int
    warning_codes: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class _WorkbookSignature:
    """记录翻译不允许改变的 workbook 结构与公式。"""

    sheets: tuple[tuple[str, int, int], ...]
    merged_ranges: tuple[tuple[int, tuple[str, ...]], ...]
    formulas: tuple[tuple[int, str, str], ...]
    tables: tuple[tuple[int, str, str], ...]


class XlsxPipeline:
    """同步执行文件进、文件出的结构保真 XLSX 翻译。"""

    document_kind = "xlsx"

    def __init__(self, translator: BatchTranslator) -> None:
        """绑定 provider-neutral translator 与格式实现。"""

        self._translator = translator
        self._extractor = XlsxExtractor()
        self._formatter = XlsxFormatter()

    def translate(
        self,
        request: TranslationRequest,
        *,
        report_progress: TranslationProgressReporter | None = None,
    ) -> TranslationResult:
        """翻译共享 request, 并生成不会覆盖源文件的输出名。"""

        language = _SAFE_LANGUAGE_RE.sub("-", request.target_language).strip("-") or "translated"
        output_path = request.output_dir / f"{request.source_path.stem}.{language}.xlsx"
        return self.translate_to(
            source_path=request.source_path,
            output_path=output_path,
            source_language=request.source_language,
            target_language=request.target_language,
            report_progress=report_progress,
        )

    def translate_to(
        self,
        *,
        source_path: str | Path,
        output_path: str | Path,
        source_language: str | None,
        target_language: str,
        report_progress: TranslationProgressReporter | None = None,
    ) -> TranslationResult:
        """把一个 workbook 翻译到显式目标路径。"""

        if report_progress is not None:
            report_progress(TranslationProgress(stage="extracting"))
        source = Path(source_path)
        output = Path(output_path)
        self._validate_paths(source, output)
        output.parent.mkdir(parents=True, exist_ok=True)
        expected_signature = _workbook_signature(source)
        segments = self._extractor.extract(source)
        total_segments = len(segments)
        if report_progress is not None:
            report_progress(TranslationProgress(stage="translating", total_segments=total_segments))

        translations: dict[tuple[int, str], str] = {}
        stats = _TranslationStats()
        self._translate_segments(
            segments,
            translations=translations,
            stats=stats,
            source_language=source_language,
            target_language=target_language,
            report_progress=report_progress,
        )
        if report_progress is not None:
            report_progress(
                TranslationProgress(
                    stage="formatting",
                    processed_segments=stats.processed_segments,
                    total_segments=total_segments,
                )
            )
        self._write_atomically(
            source,
            output,
            segments,
            translations,
            expected_signature=expected_signature,
        )
        return TranslationResult(
            output_path=output,
            document_kind="xlsx",
            translated_segments=stats.translated_segments,
            fallback_segments=stats.fallback_segments,
            warning_codes=tuple(sorted(stats.warning_codes)),
        )

    def _translate_segments(
        self,
        segments: Sequence[XlsxCellSegment],
        *,
        translations: dict[tuple[int, str], str],
        stats: _TranslationStats,
        source_language: str | None,
        target_language: str,
        report_progress: TranslationProgressReporter | None,
    ) -> None:
        """按 token 分组并有界并发, 为每个 cell 留下安全结果。"""

        groups = group_by_token_limit(segments, lambda segment: segment.translation_text)

        def translate_group(group: tuple[XlsxCellSegment, ...]) -> _GroupTranslation:
            """让一个 worker 完成 provider 调用与安全 fallback。"""

            return self._translate_group(
                group,
                source_language=source_language,
                target_language=target_language,
            )

        def commit_group(
            _group_index: int,
            group: tuple[XlsxCellSegment, ...],
            outcome: BatchFanoutOutcome[_GroupTranslation],
        ) -> None:
            """按原 group 顺序写回结果并推进真实进度。"""

            group_result = outcome.value
            if group_result is None:
                group_result = _GroupTranslation(
                    items=tuple((segment.key, segment.original_text) for segment in group),
                    translated_segments=0,
                    fallback_segments=len(group),
                    warning_codes=("xlsx_cell_fallback",),
                )
            translations.update(group_result.items)
            stats.translated_segments += group_result.translated_segments
            stats.fallback_segments += group_result.fallback_segments
            stats.warning_codes.update(group_result.warning_codes)
            stats.processed_segments += len(group)
            if report_progress is not None:
                report_progress(
                    TranslationProgress(
                        stage="translating",
                        processed_segments=stats.processed_segments,
                        total_segments=len(segments),
                    )
                )

        run_batch_fanout(
            groups,
            translate_group,
            max_concurrency=translator_per_job_concurrency(self._translator),
            on_group_settled=commit_group,
        )

    def _translate_group(
        self,
        group: Sequence[XlsxCellSegment],
        *,
        source_language: str | None,
        target_language: str,
    ) -> _GroupTranslation:
        """接受 marker 与唯一 span 都完整的译文, 其余 cell 回退原文。"""

        result: TranslationBatchResult | None
        try:
            result = self._translator.translate_batch(
                texts=[segment.translation_text for segment in group],
                source_language=source_language,
                target_language=target_language,
                format_hint="xlsx",
            )
        except Exception:
            result = None
        candidates = _index_batch_result(result, len(group))
        items: list[tuple[tuple[int, str], str]] = []
        translated_segments = 0
        fallback_segments = 0
        for index, segment in enumerate(group):
            translated = _accepted_cell_text(segment, candidates.get(index))
            if translated is None:
                translated = segment.original_text
                fallback_segments += 1
            else:
                translated_segments += 1
            items.append((segment.key, translated))
        warnings = ("xlsx_cell_fallback",) if fallback_segments else ()
        return _GroupTranslation(
            items=tuple(items),
            translated_segments=translated_segments,
            fallback_segments=fallback_segments,
            warning_codes=warnings,
        )

    def _write_atomically(
        self,
        source: Path,
        output: Path,
        segments: Sequence[XlsxCellSegment],
        translations: dict[tuple[int, str], str],
        *,
        expected_signature: _WorkbookSignature,
    ) -> None:
        """在同目录临时文件完成写回和结构校验后再发布。"""

        descriptor, temp_name = tempfile.mkstemp(
            dir=output.parent,
            prefix=f".{output.name}.",
            suffix=".tmp.xlsx",
        )
        os.close(descriptor)
        temporary = Path(temp_name)
        try:
            self._formatter.apply(source, temporary, segments, translations)
            with temporary.open("r+b") as handle:
                handle.flush()
                os.fsync(handle.fileno())
            if _workbook_signature(temporary) != expected_signature:
                raise ValueError("xlsx_structure_signature_changed")
            os.replace(temporary, output)
        except Exception:
            temporary.unlink(missing_ok=True)
            raise

    @staticmethod
    def _validate_paths(source: Path, output: Path) -> None:
        """拒绝缺失、格式错误和覆盖源文件的路径。"""

        if not source.is_file():
            raise FileNotFoundError(source)
        if source.suffix.lower() != ".xlsx":
            raise ValueError("document_kind_mismatch")
        if source.resolve() == output.resolve():
            raise ValueError("output_would_overwrite_source")


def _accepted_cell_text(segment: XlsxCellSegment, candidate: str | None) -> str | None:
    """从 provider 候选中提取当前 cell 的单一 span 文本。"""

    if candidate is None:
        return None
    value = candidate.strip("\r\n")
    marker = f"[{segment.marker}]"
    if not value.startswith(marker):
        return None
    match = _SPAN_RE.fullmatch(value[len(marker) :])
    if match is None:
        return None
    translated = unescape(match.group(1))
    return translated if translated.strip() else None


def _index_batch_result(
    result: TranslationBatchResult | None,
    expected_count: int,
) -> dict[int, str]:
    """只保留范围内且恰好出现一次的 provider 结果 index。"""

    if result is None:
        return {}
    counts = Counter(item.index for item in result.items)
    return {
        item.index: item.text
        for item in result.items
        if 0 <= item.index < expected_count and counts[item.index] == 1
    }


def _workbook_signature(path: Path) -> _WorkbookSignature:
    """读取 sheet、合并区、公式与 Table 范围, 忽略允许变化的 cell 文本。"""

    workbook = load_workbook(path, data_only=False, read_only=False)
    sheets: list[tuple[str, int, int]] = []
    merged_ranges: list[tuple[int, tuple[str, ...]]] = []
    formulas: list[tuple[int, str, str]] = []
    tables: list[tuple[int, str, str]] = []
    try:
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            sheets.append((worksheet.title, worksheet.max_row or 0, worksheet.max_column or 0))
            merged_ranges.append(
                (sheet_index, tuple(sorted(str(item) for item in worksheet.merged_cells.ranges)))
            )
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        formulas.append((sheet_index, cell.coordinate, str(cell.value)))
            for table in worksheet.tables.values():
                tables.append((sheet_index, str(table.name), str(table.ref)))
    finally:
        workbook.close()
    return _WorkbookSignature(
        sheets=tuple(sheets),
        merged_ranges=tuple(merged_ranges),
        formulas=tuple(formulas),
        tables=tuple(sorted(tables)),
    )
