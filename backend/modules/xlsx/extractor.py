"""从 XLSX 中提取可翻译字符串, 同时跳过高风险结构内容。"""

import re
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, range_boundaries

from .entities import XlsxCellSegment, XlsxTableHeaderRef


class XlsxExtractor:
    """遍历 workbook, 并只暴露不会破坏公式和引用的字符串 cell。"""

    _CODE_PATTERNS = (
        re.compile(r"^(?:https?://|www\.)", re.IGNORECASE),
        re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$"),
        re.compile(r"^[A-Za-z]:\\"),
        re.compile(r"^/[A-Za-z0-9._/\-]+$"),
        re.compile(
            r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[1-5][0-9a-fA-F]{3}-"
            r"[89abAB][0-9a-fA-F]{3}-[0-9a-fA-F]{12}$"
        ),
    )

    def extract(self, file_path: str | Path) -> tuple[XlsxCellSegment, ...]:
        """按 sheet、row、column 顺序提取可翻译单元格。"""

        workbook = load_workbook(file_path, data_only=False)
        segments: list[XlsxCellSegment] = []
        try:
            for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
                table_header_map = self._build_table_header_map(worksheet)
                merged_skip_refs = self._build_merged_skip_refs(worksheet)
                for row_index in range(1, (worksheet.max_row or 0) + 1):
                    for col_index in range(1, (worksheet.max_column or 0) + 1):
                        cell_ref = f"{get_column_letter(col_index)}{row_index}"
                        if cell_ref in merged_skip_refs:
                            continue
                        cell = worksheet.cell(row=row_index, column=col_index)
                        if not isinstance(cell.value, str) or not self._is_translatable(cell):
                            continue
                        segments.append(
                            XlsxCellSegment(
                                sheet_index=sheet_index,
                                sheet_name=worksheet.title,
                                row_index=row_index,
                                col_index=col_index,
                                cell_ref=cell_ref,
                                original_text=cell.value,
                                table_headers=tuple(table_header_map.get(cell_ref, ())),
                            )
                        )
        finally:
            workbook.close()
        return tuple(segments)

    def _build_table_header_map(self, worksheet: object) -> dict[str, list[XlsxTableHeaderRef]]:
        """把 Table header cell 映射到必须同步更新的列 metadata。"""

        header_map: dict[str, list[XlsxTableHeaderRef]] = defaultdict(list)
        for table in worksheet.tables.values():
            table_ref = getattr(table, "ref", None)
            table_name = getattr(table, "name", "")
            if not table_ref or not table_name:
                continue
            try:
                min_col, min_row, max_col, _ = range_boundaries(table_ref)
            except ValueError:
                continue
            for column_offset, col_index in enumerate(range(min_col, max_col + 1), start=1):
                cell_ref = f"{get_column_letter(col_index)}{min_row}"
                header_map[cell_ref].append(
                    XlsxTableHeaderRef(table_name=table_name, column_index=column_offset)
                )
        return header_map

    @staticmethod
    def _build_merged_skip_refs(worksheet: object) -> set[str]:
        """返回合并区域中除左上角主 cell 之外的所有坐标。"""

        skip_refs: set[str] = set()
        for merged in worksheet.merged_cells.ranges:
            master_ref = merged.start_cell.coordinate
            for row_index, col_index in merged.cells:
                cell_ref = f"{get_column_letter(col_index)}{row_index}"
                if cell_ref != master_ref:
                    skip_refs.add(cell_ref)
        return skip_refs

    def _is_translatable(self, cell: Cell) -> bool:
        """拒绝公式、结构标识符以及没有自然语言字符的文本。"""

        text = str(cell.value)
        normalized = text.strip()
        if cell.data_type == "f" or not normalized or normalized.startswith("="):
            return False
        if any(pattern.match(normalized) for pattern in self._CODE_PATTERNS):
            return False
        if re.fullmatch(r"[A-Za-z0-9._:/\\\-]+", normalized):
            has_letter = re.search(r"[A-Za-z]", normalized) is not None
            has_digit = re.search(r"\d", normalized) is not None
            has_separator = re.search(r"[._:/\\\-]", normalized) is not None
            if has_letter and has_digit and has_separator:
                return False
        return not all(
            character.isspace() or unicodedata.category(character)[:1] in {"N", "P", "S"}
            for character in normalized
        )
