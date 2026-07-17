"""验证 XLSX pipeline 的筛选、翻译、回退与结构保护。"""

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from modules.translation.contracts import TranslationBatchItem, TranslationBatchResult
from modules.xlsx import XlsxPipeline


class XlsxTranslator:
    """把每个 span 内容加上前缀, 并记录实际模型调用。"""

    def __init__(self, *, break_cell: str | None = None) -> None:
        """保存需要故意破坏 marker 的 cell。"""

        self.break_cell = break_cell
        self.calls: list[tuple[str, ...]] = []

    def translate_batch(
        self,
        *,
        texts: Sequence[str],
        source_language: str | None,
        target_language: str,
        format_hint: str,
        read_only_context: Sequence[str] = (),
        repair_candidates: Sequence[str] = (),
    ) -> TranslationBatchResult:
        """保持 index 对齐, 只对指定 cell 返回坏结构。"""

        del source_language, target_language, read_only_context, repair_candidates
        assert format_hint == "xlsx"
        captured = tuple(texts)
        self.calls.append(captured)
        items: list[TranslationBatchItem] = []
        for index, text in enumerate(captured):
            if self.break_cell and f"[CELL_{self.break_cell}]" in text:
                translated = "broken"
            else:
                translated = text.replace("<span>", "<span>EN_")
            items.append(TranslationBatchItem(index=index, text=translated))
        return TranslationBatchResult(items=tuple(items))


def _sample_workbook(path: Path) -> None:
    """创建含 Table、公式、标识符与合并单元格的 workbook。"""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sales"
    worksheet["A1"] = "产品"
    worksheet["B1"] = "数量"
    worksheet["A2"] = "苹果"
    worksheet["B2"] = 10
    worksheet["A3"] = "SKU-001"
    worksheet["B3"] = "=SUM(B2,1)"
    worksheet.merge_cells("C1:D1")
    worksheet["C1"] = "合并标题"
    table = Table(displayName="SalesTable", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    workbook.save(path)
    workbook.close()


def test_xlsx_pipeline_translates_safe_cells_and_preserves_structure(tmp_path: Path) -> None:
    """字符串 cell 可翻译, 公式、标识符、合并区与 Table contract 必须保留。"""

    source = tmp_path / "source.xlsx"
    output = tmp_path / "translated.xlsx"
    _sample_workbook(source)
    original = source.read_bytes()
    translator = XlsxTranslator()

    result = XlsxPipeline(translator).translate_to(
        source_path=source,
        output_path=output,
        source_language="zh-CN",
        target_language="en",
    )

    assert source.read_bytes() == original
    assert result.translated_segments == 4
    assert result.fallback_segments == 0
    assert result.warning_codes == ()
    assert len(translator.calls) == 1
    workbook = load_workbook(output, data_only=False)
    try:
        worksheet = workbook["Sales"]
        assert worksheet["A1"].value == "EN_产品"
        assert worksheet["B1"].value == "EN_数量"
        assert worksheet["A2"].value == "EN_苹果"
        assert worksheet["A3"].value == "SKU-001"
        assert worksheet["B3"].value == "=SUM(B2,1)"
        assert worksheet["C1"].value == "EN_合并标题"
        assert worksheet["D1"].value is None
        assert next(iter(worksheet.merged_cells.ranges)).coord == "C1:D1"
        table = worksheet.tables["SalesTable"]
        assert [column.name for column in table.tableColumns] == ["EN_产品", "EN_数量"]
    finally:
        workbook.close()


def test_xlsx_pipeline_falls_back_only_the_cell_with_broken_marker(tmp_path: Path) -> None:
    """单个坏候选只能让自己的 cell 回退, 不能污染其他 cell。"""

    source = tmp_path / "source.xlsx"
    output = tmp_path / "translated.xlsx"
    _sample_workbook(source)

    result = XlsxPipeline(XlsxTranslator(break_cell="A2")).translate_to(
        source_path=source,
        output_path=output,
        source_language="zh-CN",
        target_language="en",
    )

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook["Sales"]["A1"].value == "EN_产品"
        assert workbook["Sales"]["A2"].value == "苹果"
    finally:
        workbook.close()
    assert result.translated_segments == 3
    assert result.fallback_segments == 1
    assert result.warning_codes == ("xlsx_cell_fallback",)
